# -*- coding: utf-8 -*-
from __future__ import print_function
from __future__ import unicode_literals
import os
import re
import io
import sys
import json
from collections import defaultdict

try:
    from openpyxl import Workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

def get_src_context_with_linenum(src_context, start_line):
    """为每行代码添加行号"""
    updated_context = []
    for idx, ctx_line in enumerate(src_context):
        ctx_line = ctx_line.strip() + "//##" + str(start_line + idx)
        updated_context.append(ctx_line)
    return updated_context

def classify_line_type(line):
    """根据目标行代码进行类型分类"""
    line = line.strip()
    if not line:
        return 'Other'

    # 返回语句
    if line.startswith('return '):
        return 'Return Statement'

    # 循环语句
    if line.startswith('for (') or line.startswith('while ('):
        return 'Loop Statement'

    # 结构体/联合体/枚举声明
    if any(kw in line for kw in ['struct ', 'union ', 'enum ']) and '{' in line:
        return 'Type Declaration'

    # 变量声明（带赋值）
    if re.match(r'(const\s+)?(struct|int|char|long|void|unsigned|bool|float|double)\s+\w+.*=', line):
        return 'Variable Declaration with Assignment'

    # 纯变量声明
    if re.match(r'(const\s+)?(struct|int|char|long|void|unsigned|bool|float|double)\s+\w+', line):
        return 'Variable Declaration'

    # 赋值语句（不是声明）
    if '=' in line and '==' not in line and not re.match(r'(const\s+)?(struct|int|char|long|void|unsigned|bool|float|double)', line):
        return 'Assignment Statement'

    # 函数调用（带赋值）
    if re.search(r'\w+\s*=\s*\w+\s*\([^)]*\)', line):
        return 'Function Call with Assignment'

    # 函数调用（无赋值）
    if re.search(r'^\s*\w+\s*\([^)]*\)', line):
        return 'Function Call'

    # 数组/指针操作
    if '[' in line or ']' in line:
        return 'Array/Pointer Access'

    return 'Other'

def find_function_end_line(filepath, func_start_line):
    """找到函数的结束行（配对的闭花括号）"""
    try:
        with io.open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
            lines = f.readlines()

        if func_start_line < 1 or func_start_line > len(lines):
            return None

        # 从函数定义行开始，找到第一个 {
        brace_count = 0
        found_opening = False

        for i in range(func_start_line - 1, len(lines)):
            line = lines[i]

            # 计算括号
            for char in line:
                if char == '{':
                    brace_count += 1
                    found_opening = True
                elif char == '}':
                    if found_opening:
                        brace_count -= 1
                        if brace_count == 0:
                            return i + 1  # 返回1-based行号

        return None
    except:
        return None


def find_function_definition_line(filepath, func_name, start_from_line=None):
    """
    查找函数定义的行号
    如果提供了 start_from_line，则从该行向上搜索（推荐方式）
    否则从文件开始向下搜索
    """
    try:
        with io.open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
            lines = f.readlines()

        pattern = re.compile(r'\b' + re.escape(func_name) + r'\s*\(')

        # 模式：识别函数定义（返回类型 + 函数名 + ( + 可能的参数）
        def is_function_definition(line_idx):
            """检查这一行是否是函数定义"""
            line = lines[line_idx]
            stripped = line.strip()

            # 排除注释行
            if stripped.startswith('//') or stripped.startswith('/*') or stripped.startswith('*'):
                return False

            # 检查是否有函数名和括号
            if not pattern.search(line):
                return False

            # 检查当前行是否有函数定义特征（返回类型关键字）
            if any(kw in stripped for kw in ['void ', 'int ', 'bool ', 'char ', 'long ', 'struct ', 'static ', 'inline ', 'unsigned ', 'const ', '*']):
                return True

            # 如果当前行没有返回类型，检查前面的几行（处理多行函数定义）
            # 向上查找最多10行，寻找返回类型关键字
            for i in range(max(0, line_idx - 10), line_idx):
                prev_line = lines[i].strip()
                if any(kw in prev_line for kw in ['int ', 'void ', 'bool ', 'char ', 'long ', 'struct ', 'static ', 'inline ', 'unsigned ', 'const ', '*']):
                    # 确保这几行之间没有分号或其他语句结束符
                    between_lines = '\n'.join(lines[i:line_idx+1])
                    if ';' not in between_lines.split('{')[0]:  # 不考虑{}中的分号
                        return True

            return False

        # 策略 1：如果提供了 start_from_line，从那里向上搜索（推荐）
        if start_from_line is not None:
            start_idx = min(start_from_line - 1, len(lines) - 1)
            for i in range(start_idx, max(-1, start_idx - 500), -1):  # 向上搜索最多 500 行
                if is_function_definition(i):
                    return i + 1

        # 策略 2：从文件开始向下搜索
        for lineno in range(len(lines)):
            if is_function_definition(lineno):
                return lineno + 1

        return None
    except:
        return None

def is_conditional_statement(line):
    """检查一行是否是条件语句（if、else if、else、while、for等）"""
    line_stripped = line.strip()
    if not line_stripped:
        return False

    # 检查是否是条件语句关键字开头
    keywords = ['if ', 'else if ', 'else', 'while ', 'for ', 'switch ', 'case ']
    for kw in keywords:
        if line_stripped.startswith(kw):
            return True

    return False

def references_any_arg(line, args):
    """检查一行是否引用了任何参数"""
    if not args:
        return True

    for arg in args:
        # 使用word boundary来匹配参数名
        if re.search(r'\b' + re.escape(arg) + r'\b', line):
            return True
    return False

def filter_lines_with_callsite_args(code_lines, callsite_code, start_line):
    """
    从代码行列表中过滤出包含 Callsite 参数引用的行
    逻辑：
    1. 保留 callsite 行本身
    2. 仅删除不引用 Callsite Argument 的条件语句（if、else、while、for等）
    3. 其他行全部保留（无论是否引用参数）
    4. 为每行添加原始行号
    """
    if not callsite_code:
        return get_src_context_with_linenum(code_lines, start_line)

    # 从 callsite 代码中提取参数
    # 查找函数调用中的参数部分
    match = re.search(r'\w+\s*\(([^)]*)\)', callsite_code)
    if not match:
        return get_src_context_with_linenum(code_lines, start_line)

    args_str = match.group(1)
    if not args_str.strip():
        return get_src_context_with_linenum(code_lines, start_line)

    # 解析参数列表（简单分割）
    args = [arg.strip() for arg in args_str.split(',')]
    # 只保留有效的参数标识符（去掉空字符串和复杂表达式的一部分）
    args = [arg for arg in args if arg and not any(c in arg for c in ['(', ')', '[', ']'])]

    filtered_lines = []
    for idx, line in enumerate(code_lines):
        line_stripped = line.strip()
        original_line_num = start_line + idx

        # 保留callsite行本身
        if callsite_code in line_stripped:
            filtered_lines.append(line_stripped + "//##" + str(original_line_num))
        # 对于条件语句，仅当不引用参数时才删除
        elif is_conditional_statement(line):
            if references_any_arg(line, args):
                # 引用参数的条件语句：保留
                filtered_lines.append(line_stripped + "//##" + str(original_line_num))
            # 不引用参数的条件语句：删除（不保留）
        # 对于非条件语句，全部保留
        else:
            filtered_lines.append(line_stripped + "//##" + str(original_line_num))

    return filtered_lines


def _generate_exclude_list(xlsx_file, output_list_file):
    """
    从 xlsx 文件生成排除列表（简单的文本格式）
    格式：target_file|target_line|callsite_line
    """
    try:
        from openpyxl import load_workbook
        wb = load_workbook(xlsx_file)
        ws = wb.active

        with io.open(output_list_file, 'w', encoding='utf-8') as f:
            for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
                if row_idx == 1:  # Skip header
                    continue

                if row and len(row) >= 8:
                    try:
                        target_file = str(row[0]).strip() if row[0] else ''
                        target_line = str(row[1]).strip() if row[1] else ''
                        callsite_line = str(row[7]).strip() if row[7] else ''

                        if target_file and target_line and callsite_line:
                            line = u"{}|{}|{}\n".format(target_file, target_line, callsite_line)
                            f.write(line)
                    except:
                        pass

        return output_list_file
    except:
        return None


def _read_exclude_list(list_file):
    """
    从排除列表文件读取记录
    """
    records_set = set()

    if not os.path.exists(list_file):
        return records_set

    try:
        with io.open(list_file, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if line:
                    parts = line.split('|')
                    if len(parts) == 3:
                        record_id = tuple(parts)
                        records_set.add(record_id)

        print("[INFO] Loaded {} excluded records from list".format(len(records_set)))
        return records_set
    except:
        print("[WARNING] Failed to read exclude list")
        return records_set


def _read_xlsx_records(xlsx_file):
    """
    从 Excel 文件读取记录，返回唯一标识集合
    标识：(target_func_filepath, target_line, callsite_line)
    """
    records_set = set()

    if not os.path.exists(xlsx_file):
        return records_set

    # 直接用 openpyxl 读取，不处理错误信息以避免编码问题
    try:
        from openpyxl import load_workbook
        wb = load_workbook(xlsx_file, data_only=True)
        ws = wb.active

        # 跳过 header 行
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            if row_idx == 1:  # Skip header
                continue

            if row and len(row) >= 8:
                try:
                    target_file = str(row[0]).strip() if row[0] else ''
                    target_line = str(row[1]).strip() if row[1] else ''
                    callsite_line = str(row[7]).strip() if row[7] else ''

                    if target_file and target_line and callsite_line:
                        record_id = (target_file, target_line, callsite_line)
                        records_set.add(record_id)
                except:
                    pass

        return records_set

    except:
        return records_set


def extract_pattern_type_2_data(output_suffix='1', exclude_xlsx=None, excluded_records=None):
    """
    从unchecked_locations_report.json中提取pattern_type为2的数据，
    平均分布在不同的内核子系统中，输出为mc_cross_{suffix}.list格式

    参数：
        output_suffix: 输出文件后缀（默认 '1'，可设置为 '2' 等）
        exclude_xlsx: 要排除的 xlsx 文件路径（已弃用，使用 excluded_records 代替）
        excluded_records: 要排除的记录集合（格式：set of (target_file, target_line, callsite_line)）
    """
    if isinstance(__file__, bytes):
        file_path = __file__.decode('utf-8')
    else:
        file_path = __file__
    project_dir = os.path.dirname(os.path.dirname(os.path.abspath(file_path)))
    if isinstance(project_dir, bytes):
        project_dir = project_dir.decode('utf-8')
    json_file = os.path.join(project_dir, u'scripts', u'unchecked_locations_report.json')
    output_file = os.path.join(project_dir, u'data', u'test_data', u'mc_cross_{}.xlsx'.format(output_suffix))
    temp_output_file = output_file + u'.tmp'
    target_count = 100

    # 确定排除的记录集合
    excluded_records_set = set()

    # 根据 output_suffix 确定要排除的文件
    if output_suffix == '2':
        # mc_cross_2: 排除 mc_cross_1
        selected_files = [u'mc_cross_1_selected.txt']
        exclude_label = 'mc_cross_1'
    elif output_suffix == '3':
        # mc_cross_3: 排除 mc_cross_1 和 mc_cross_2
        selected_files = [u'mc_cross_1_selected.txt', u'mc_cross_2_selected.txt']
        exclude_label = 'mc_cross_1 and mc_cross_2'
    else:
        selected_files = []
        exclude_label = None

    # 读取选中记录文件
    if selected_files:
        for selected_file_name in selected_files:
            selected_records_file = os.path.join(project_dir, u'data', u'test_data', selected_file_name)

            # 如果选中记录文件不存在，尝试从对应的 xlsx 生成
            if not os.path.exists(selected_records_file):
                # 从文件名推断 xlsx 文件名
                xlsx_suffix = selected_file_name.split('_')[2].split('.')[0]  # 从 mc_cross_X_selected.txt 提取 X
                mc_cross_xlsx = os.path.join(project_dir, u'data', u'test_data', u'mc_cross_{}.xlsx'.format(xlsx_suffix))
                if os.path.exists(mc_cross_xlsx):
                    _generate_exclude_list(mc_cross_xlsx, selected_records_file)

            # 尝试读取选中记录文件
            if os.path.exists(selected_records_file):
                try:
                    with io.open(selected_records_file, 'r', encoding='utf-8') as f:
                        for line in f:
                            line = line.strip()
                            if line:
                                parts = line.split('|')
                                if len(parts) == 3:
                                    excluded_records_set.add(tuple(parts))
                except:
                    pass

        if excluded_records_set:
            try:
                print("[INFO] Loaded {} records to exclude from {}".format(len(excluded_records_set), exclude_label))
            except:
                pass

    if excluded_records:
        excluded_records_set = excluded_records
    elif exclude_xlsx:
        excluded_records_set = _read_xlsx_records(exclude_xlsx)

    try:
        print("[EXTRACT] Reading {}".format(json_file))
    except:
        print("[EXTRACT] Reading JSON file...")
    if not os.path.exists(json_file):
        try:
            print("[ERROR] File not found: {}".format(json_file))
        except:
            print("[ERROR] File not found")
        return

    # 读取JSON文件
    with io.open(json_file, 'r', encoding='utf-8') as f:
        all_data = json.load(f)

    # 按pattern_type == 2进行过滤，并按子系统和文件分组
    data_by_subsystem_file = defaultdict(lambda: defaultdict(list))
    pattern_type_2_data = [d for d in all_data if d.get('pattern_type') == 2]

    print("[EXTRACT] Found {} records with pattern_type=2".format(len(pattern_type_2_data)))

    # 如果有排除集合，过滤出不在其中的记录
    if excluded_records_set:
        original_count = len(pattern_type_2_data)
        pattern_type_2_data = [
            d for d in pattern_type_2_data
            if (d.get('target_func_filepath', ''), str(d.get('target_line', '')), str(d.get('callsite_line', '')))
            not in excluded_records_set
        ]
        print("[EXTRACT] After excluding: {} records (filtered out {} duplicates)".format(
            len(pattern_type_2_data), original_count - len(pattern_type_2_data)))

    for item in pattern_type_2_data:
        subsystem = item.get('subsystem', 'unknown')
        filepath = item.get('filepath', 'unknown')
        data_by_subsystem_file[subsystem][filepath].append(item)

    print("[EXTRACT] Distribution by subsystem and file:")
    subsystems = sorted(data_by_subsystem_file.keys())
    for subsystem in subsystems:
        file_count = len(data_by_subsystem_file[subsystem])
        total_records = sum(len(items) for items in data_by_subsystem_file[subsystem].values())
        print("  {}: {} files, {} records".format(subsystem, file_count, total_records))

    # 全局采样：轮询各子系统，轮询各文件，确保总数恰好为target_count
    selected_data = []
    global_samples_per_file = defaultdict(int)
    subsystem_file_index = {}

    print("\n[EXTRACT] Extracting {} records (distributing across subsystems and files):".format(target_count))
    print("[DEBUG] Subsystems (sorted): {}".format(subsystems))

    # 保存选中的记录列表，供后续的mc_cross_N+1使用
    selected_records_file = os.path.join(project_dir, u'data', u'test_data', u'mc_cross_{}_selected.txt'.format(output_suffix))

    # 初始化每个子系统的文件列表和采样位置
    for subsystem in subsystems:
        files = sorted(data_by_subsystem_file[subsystem].keys())
        subsystem_file_index[subsystem] = {f: 0 for f in files}

    # 轮询采样：循环遍历子系统和文件，直到达到目标数量或所有记录用完
    subsystem_idx = 0
    all_exhausted = False  # 标志：所有子系统的记录都已用完

    while len(selected_data) < target_count and not all_exhausted:
        if subsystem_idx >= len(subsystems):
            subsystem_idx = 0

        subsystem = subsystems[subsystem_idx]
        files_in_subsystem = sorted(data_by_subsystem_file[subsystem].keys())

        # 在该子系统中轮询选择文件
        file_found = False
        for _ in range(len(files_in_subsystem)):
            # 找到有可用样本的文件
            for filepath in files_in_subsystem:
                records_in_file = data_by_subsystem_file[subsystem][filepath]
                current_idx = subsystem_file_index[subsystem][filepath]

                if current_idx < len(records_in_file):
                    # 从该文件提取一条样本
                    record = records_in_file[current_idx]
                    selected_data.append(record)
                    subsystem_file_index[subsystem][filepath] += 1
                    global_samples_per_file[filepath] += 1
                    file_found = True
                    break

            if file_found or len(selected_data) >= target_count:
                break

        subsystem_idx += 1

        # 检查是否所有子系统的记录都已用完
        all_exhausted = all(
            all(subsystem_file_index[sub][f] >= len(data_by_subsystem_file[sub][f])
                for f in sorted(data_by_subsystem_file[sub].keys()))
            for sub in subsystems
        )

        if len(selected_data) >= target_count:
            break

    # 截断到目标数量
    selected_data = selected_data[:target_count]

    # 如果实际获得的记录数少于目标数量，提示警告
    if len(selected_data) < target_count:
        try:
            print("[WARNING] Only {} records available (target was {}), using all available records".format(
                len(selected_data), target_count))
        except:
            pass

    # 生成采样摘要用于验证（确保每次一致）
    try:
        sample_count = min(5, len(selected_data))
        print("[DEBUG] Sample verification (first {} records):".format(sample_count))
        for idx in range(sample_count):
            record = selected_data[idx]
            subsystem = record.get('subsystem', 'unknown')
            try:
                print("  Record {}: {}".format(idx + 1, subsystem))
            except:
                print("  Record {}".format(idx + 1))
    except:
        pass

    # 统计结果
    subsystem_counts = defaultdict(int)
    subsystem_files = defaultdict(set)
    for record in selected_data:
        subsystem = record.get('subsystem', 'unknown')
        subsystem_counts[subsystem] += 1
        filepath = record.get('filepath', 'unknown')
        subsystem_files[subsystem].add(filepath)

    for subsystem in sorted(subsystem_counts.keys()):
        print("  {}: {} records from {} files".format(
            subsystem, subsystem_counts[subsystem], len(subsystem_files[subsystem])))

    print("[EXTRACT] Total selected: {} records".format(len(selected_data)))

    # 如果是生成 mc_cross_1，保存选中的记录到文件供 mc_cross_2 使用
    if selected_records_file:
        try:
            with io.open(selected_records_file, 'w', encoding='utf-8') as f:
                for record in selected_data:
                    target_file = record.get('target_func_filepath', '')
                    target_line = str(record.get('target_line', ''))
                    callsite_line = str(record.get('callsite_line', ''))
                    if target_file and target_line and callsite_line:
                        line = u"{}|{}|{}\n".format(target_file, target_line, callsite_line)
                        f.write(line)
        except:
            pass

    # 转换路径格式（从完整路径提取相对部分）
    def extract_relative_path(filepath):
        """从完整路径中提取linux-*相对部分"""
        # 查找 linux- 的位置
        match = re.search(r'(linux-[\w\./-]+)', filepath)
        if match:
            return match.group(1).replace('\\', '/')
        return filepath

    # 输出为 Excel 格式或 txt 格式
    try:
        print("[EXTRACT] Writing to output file...")
    except:
        print("[EXTRACT] Writing to output file...")

    valid_records = 0

    # 尝试使用 openpyxl 生成 Excel 文件
    if OPENPYXL_AVAILABLE:
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = u'mc_cross'

            # 写入 header
            headers = [
                u'Target File Path',
                u'Targeted Line Number',
                u'Targeted Line Code',
                u'Target Function',
                u'Caller Function',
                u'Callsite',
                u'Caller File Path',
                u'Callsite Line Number'
            ]
            ws.append(headers)

            # 写入数据
            for item in selected_data:
                target_file = item.get('target_func_filepath', '')
                target_line = str(item.get('target_line', ''))
                target_code = str(item.get('target_statement', ''))
                target_func = str(item.get('target_func', ''))
                caller_func = str(item.get('caller_func', ''))
                callsite = str(item.get('callsite_statement', ''))
                caller_file = item.get('filepath', '')
                callsite_line = str(item.get('callsite_line', ''))

                # 只写入路径有效的记录
                if target_file.strip() and caller_file.strip():
                    ws.append([
                        target_file,
                        target_line,
                        target_code,
                        target_func,
                        caller_func,
                        callsite,
                        caller_file,
                        callsite_line
                    ])
                    valid_records += 1

            # 调整列宽
            ws.column_dimensions['A'].width = 30  # Target File Path
            ws.column_dimensions['B'].width = 15  # Targeted Line Number
            ws.column_dimensions['C'].width = 25  # Targeted Line Code
            ws.column_dimensions['D'].width = 20  # Target Function
            ws.column_dimensions['E'].width = 20  # Caller Function
            ws.column_dimensions['F'].width = 30  # Callsite
            ws.column_dimensions['G'].width = 30  # Caller File Path
            ws.column_dimensions['H'].width = 18  # Callsite Line Number

            # 保存文件到临时位置，然后重命名
            wb.save(temp_output_file)
            # 删除旧文件并重命名
            try:
                if os.path.exists(output_file):
                    os.remove(output_file)
            except:
                pass
            os.rename(temp_output_file, output_file)

        except Exception as e:
            try:
                print("[WARNING] Failed to save as Excel: {}".format(str(e)))
            except:
                print("[WARNING] Failed to save as Excel")
            # 回退到txt格式
            _save_as_txt = True
    else:
        _save_as_txt = True

    # 如果不使用Excel或Excel保存失败，则保存为txt格式
    if '_save_as_txt' in locals() and _save_as_txt:
        output_txt = output_file.replace(u'.xlsx', u'.list')
        try:
            print("[EXTRACT] Saving as text format instead...")
        except:
            pass

        with io.open(output_txt, 'w', encoding='utf-8') as f:
            # 写入header
            header = u"Target File Path\tTargeted Line Number\tTargeted Line Code\tTarget Function\tCaller Function\tCallsite\tCaller File Path\tCallsite Line Number\n"
            f.write(header)

            # 写入数据
            valid_records = 0
            for item in selected_data:
                target_file = item.get('target_func_filepath', '')
                target_line = str(item.get('target_line', ''))
                target_code = str(item.get('target_statement', ''))
                target_func = str(item.get('target_func', ''))
                caller_func = str(item.get('caller_func', ''))
                callsite = str(item.get('callsite_statement', ''))
                caller_file = item.get('filepath', '')
                callsite_line = str(item.get('callsite_line', ''))

                # 只写入路径有效的记录
                if target_file.strip() and caller_file.strip():
                    line = u"{}\t{}\t{}\t{}\t{}\t{}\t{}\t{}\n".format(
                        target_file,
                        target_line,
                        target_code,
                        target_func,
                        caller_func,
                        callsite,
                        caller_file,
                        callsite_line
                    )
                    f.write(line)
                    valid_records += 1
        output_file = output_txt

    try:
        print("[SUCCESS] Extracted {} records to {} ({} valid records)".format(len(selected_data), output_file, valid_records))
    except:
        print("[SUCCESS] Extracted {} records ({} valid records)".format(len(selected_data), valid_records))

    # 按子系统统计输出
    print("\n[SUMMARY] Distribution in output:")
    subsystem_dist = defaultdict(int)
    for item in selected_data:
        subsystem = item.get('subsystem', 'unknown')
        subsystem_dist[subsystem] += 1

    for subsystem in sorted(subsystem_dist.keys()):
        print("  {}: {} records".format(subsystem, subsystem_dist[subsystem]))


def process_cross_list(input_file=None, output_prefix='mc_cross'):
    """
    处理 cross.list 文件的通用函数

    参数：
        input_file: 输入文件路径（如果为None，则使用默认的 mc_cross.list 或 sc_cross.list）
        output_prefix: 输出文件的前缀（默认 'mc_cross'，可设置为 'sc_cross' 等）
    """
    if isinstance(__file__, bytes):
        file_path = __file__.decode('utf-8')
    else:
        file_path = __file__
    project_dir = os.path.dirname(os.path.dirname(os.path.abspath(file_path)))
    if isinstance(project_dir, bytes):
        project_dir = project_dir.decode('utf-8')

    # 确定输入文件路径
    if input_file is None:
        # 如果没有指定输入文件，则根据 output_prefix 确定
        # 如果 output_prefix 中包含 'sc'，则使用 sc_cross.list，否则使用 mc_cross.list
        if 'sc' in output_prefix.lower():
            default_input_name = 'sc_cross.list'
        else:
            default_input_name = 'mc_cross.list'
        input_file = os.path.join(project_dir, u'data', u'test_data', default_input_name)
    elif not os.path.isabs(input_file):
        # 如果是相对路径，则相对于 project_dir
        input_file = os.path.join(project_dir, input_file)

    output_full_ctx = os.path.join(project_dir, u'data', u'test_data', u'{}_classify_full_contexts.list'.format(output_prefix))
    output_classify = os.path.join(project_dir, u'data', u'test_data', u'{}_classify.list'.format(output_prefix))
    kernel_code_dir = os.path.join(project_dir, u'data', u'kernel-code')

    results_full_ctx = []
    results_classify = []

    # 提取输入文件名用于显示
    try:
        input_filename = os.path.basename(input_file)
        print("[START] Processing {} with Target Function".format(input_filename))
        print("[INFO] Input file: {}".format(input_file))
        print("[INFO] Output prefix: {}".format(output_prefix))
    except:
        # 处理编码问题
        print("[START] Processing cross list file with Target Function")
        print("[INFO] Output prefix: {}".format(output_prefix))

    if not os.path.exists(input_file):
        try:
            print("[ERROR] Input file not found: {}".format(input_file))
        except:
            print("[ERROR] Input file not found (path encoding issue)")
        return

    with io.open(input_file, 'r', encoding='utf-8') as f:
        lines = f.readlines()[1:]  # Skip header

    # 去重逻辑：检测并移除重复行
    seen_lines = {}  # {line_content: first_occurrence_idx}
    duplicate_samples = []  # 记录重复的样本序号
    unique_lines = []  # 保存去重后的行

    for idx, line in enumerate(lines):
        if not line.strip():
            continue

        line_content = line.strip()
        if line_content in seen_lines:
            # 发现重复行
            first_idx = seen_lines[line_content]
            duplicate_samples.append((idx + 1, first_idx + 1))  # 1-based sample number
        else:
            # 第一次见到这一行
            seen_lines[line_content] = idx
            unique_lines.append((idx, line))

    # 如果有重复，打印出来
    if duplicate_samples:
        print("[WARNING] Found {} duplicate records:".format(len(duplicate_samples)))
        for dup_idx, first_idx in duplicate_samples:
            print("  Sample #{} is a duplicate of Sample #{}".format(dup_idx, first_idx))

    print("Processing {} records (after removing {} duplicates)".format(len(unique_lines), len(duplicate_samples)))

    for original_idx, line in unique_lines:
        parts = line.strip().split('\t')
        if len(parts) < 8:
            continue

        # 解析输入字段
        target_file = parts[0]
        target_line_num = int(parts[1])
        target_code = parts[2]
        target_func = parts[3]
        caller_func = parts[4]
        callsite_code = parts[5]
        caller_file = parts[6]
        callsite_line_num = int(parts[7])

        # 根据 target line code 进行分类（不是 callsite code）
        line_type = classify_line_type(target_code)

        # 目前默认所有行的 Target Parameter Annotation 都为 "Yes"
        param_annotation = "Yes"

        # 构造完整路径
        try:
            target_file_path = os.path.join(kernel_code_dir, target_file.replace('/', os.sep))
        except:
            target_file_path = kernel_code_dir + '\\' + target_file.replace('/', '\\')
        try:
            caller_file_path = os.path.join(kernel_code_dir, caller_file.replace('/', os.sep))
        except:
            caller_file_path = kernel_code_dir + '\\' + caller_file.replace('/', '\\')

        # 检查文件是否存在
        if not os.path.exists(target_file_path) or not os.path.exists(caller_file_path):
            continue


        # 提取上下文
        try:
            # 读取两个文件
            try:
                with io.open(caller_file_path, 'r', encoding='utf-8', errors='ignore') as f:
                    caller_content = f.readlines()
            except Exception as read_err:
                print("[ERROR] Sample #{}: Failed to read caller file: {}".format(original_idx + 1, str(read_err)))
                raise

            try:
                with io.open(target_file_path, 'r', encoding='utf-8', errors='ignore') as f:
                    target_content = f.readlines()
            except Exception as read_err:
                print("[ERROR] Sample #{}: Failed to read target file: {}".format(original_idx + 1, str(read_err)))
                raise

            context_list = []

            # 1. 找到caller函数定义行（从callsite向上搜索）
            caller_func_def_line = find_function_definition_line(caller_file_path, caller_func, callsite_line_num)

            # 2. 找到target函数定义行（从target_line向上搜索）
            target_func_def_line = find_function_definition_line(target_file_path, target_func, target_line_num)

            # 初始化上下文列表
            full_context_list = []
            classify_context_list = []

            # 调试：如果函数定义未找到，输出警告
            if not caller_func_def_line:
                try:
                    print("[WARNING] Sample #{}: Cannot find caller function '{}' in {}".format(
                        original_idx + 1, caller_func, os.path.basename(caller_file_path)))
                except:
                    pass
            if not target_func_def_line:
                try:
                    print("[WARNING] Sample #{}: Cannot find target function '{}' in {}".format(
                        original_idx + 1, target_func, os.path.basename(target_file_path)))
                except:
                    pass

            # 如果两行都找到了，提取上下文
            if caller_func_def_line and target_func_def_line:
                # ==================== full_contexts: caller函数入口到callsite处 + target function入口到target line ====================
                # 提取 caller 函数入口到 callsite 处
                full_caller_start = caller_func_def_line - 1
                full_caller_end = callsite_line_num
                caller_part_full = caller_content[full_caller_start:full_caller_end]
                caller_lines_full = get_src_context_with_linenum(caller_part_full, full_caller_start + 1)

                # 提取 target function入口到target line
                full_target_start = target_func_def_line - 1
                full_target_end = target_line_num
                target_part_full = target_content[full_target_start:full_target_end]
                target_lines_full = get_src_context_with_linenum(target_part_full, full_target_start + 1)

                full_context_list = caller_lines_full + target_lines_full

                # ==================== classify_contexts: 根据 param_annotation 选择范围 ====================
                if param_annotation == "Yes":
                    # caller 函数入口到 callsite 处，过滤出包含 Callsite Argument 引用的行
                    classify_caller_start = caller_func_def_line - 1
                    classify_caller_end = callsite_line_num
                    caller_part_classify = caller_content[classify_caller_start:classify_caller_end]
                    caller_lines_classify = filter_lines_with_callsite_args(caller_part_classify, callsite_code, classify_caller_start + 1)

                    # target function入口到target line，保留全部行
                    classify_target_start = target_func_def_line - 1
                    classify_target_end = target_line_num
                    target_part_classify = target_content[classify_target_start:classify_target_end]
                    target_lines_classify = get_src_context_with_linenum(target_part_classify, classify_target_start + 1)

                    classify_context_list = caller_lines_classify + target_lines_classify

            results_full_ctx.append({
                'file': target_file,
                'line': parts[1],
                'targeted_line_code': target_code,
                'caller_function': caller_func,
                'callsite': callsite_code,
                'context': full_context_list,
                'type': line_type,
                'param_annotation': param_annotation
            })

            results_classify.append({
                'file': target_file,
                'line': parts[1],
                'targeted_line_code': target_code,
                'caller_function': caller_func,
                'callsite': callsite_code,
                'context': classify_context_list,
                'type': line_type,
                'param_annotation': param_annotation
            })

        except Exception as e:
            pass

    # 保存结果
    print("\n[SAVE] {} full_context records".format(len(results_full_ctx)))
    with io.open(output_full_ctx, 'w', encoding='utf-8') as f:
        for r in results_full_ctx:
            ctx_str = str(r['context'])
            line = "{} ;; {} ;; {} ;; {} ;; {} ;; {} ;; {} ;; {}\n".format(
                '../data/kernel-code/' + r['file'],
                r['line'],
                r['targeted_line_code'],
                r['caller_function'],
                r['callsite'],
                ctx_str,
                r['type'],
                r['param_annotation']
            )
            f.write(line)

    print("[SAVE] {} classify records".format(len(results_classify)))
    with io.open(output_classify, 'w', encoding='utf-8') as f:
        for r in results_classify:
            ctx_str = str(r['context'])
            line = "{} ;; {} ;; {} ;; {} ;; {} ;; {} ;; {} ;; {}\n".format(
                '../data/kernel-code/' + r['file'],
                r['line'],
                r['targeted_line_code'],
                r['caller_function'],
                r['callsite'],
                ctx_str,
                r['type'],
                r['param_annotation']
            )
            f.write(line)

    print("\n[SUMMARY]")
    dist = {}
    for r in results_classify:
        t = r['type']
        dist[t] = dist.get(t, 0) + 1
    print("  Total: {}".format(len(results_classify)))
    for t, c in sorted(dist.items()):
        print("    {}: {}".format(t, c))


def process_mc_cross_list():
    """处理 mc_cross.list 文件"""
    process_cross_list(input_file=None, output_prefix='mc_cross')


def process_sc_cross_list():
    """处理 sc_cross.list 文件"""
    process_cross_list(input_file=None, output_prefix='sc_cross')


# 处理命令行参数
if __name__ == '__main__':
    import argparse

    parser = argparse.ArgumentParser(description='Process cross.list files and extract context')
    parser.add_argument('--extract-cross', action='store_true',
                        help='Extract pattern_type=2 data from unchecked_locations_report.json (mc_cross_1)')
    parser.add_argument('--extract-cross-2', action='store_true',
                        help='Extract remaining pattern_type=2 data (excluding mc_cross_1, generate mc_cross_2)')
    parser.add_argument('--extract-cross-3', action='store_true',
                        help='Extract remaining pattern_type=2 data (excluding mc_cross_1 and mc_cross_2, generate mc_cross_3)')
    parser.add_argument('--gen-exclude-list', action='store_true',
                        help='Generate exclude list from mc_cross_1.xlsx for mc_cross_2 extraction')
    parser.add_argument('--extract-cross-context', action='store_true',
                        help='Process mc_cross.list and generate classify results')
    parser.add_argument('--extract-sc-cross-context', action='store_true',
                        help='Process sc_cross.list and generate classify results')
    parser.add_argument('--input-file', type=str, default=None,
                        help='Custom input file path (absolute or relative to project directory)')
    parser.add_argument('--output-prefix', type=str, default=None,
                        help='Output file prefix (default: depends on the data type)')

    args = parser.parse_args()

    if args.extract_cross:
        # 从unchecked_locations_report.json提取pattern_type=2的数据，生成mc_cross_1.xlsx
        extract_pattern_type_2_data(output_suffix='1')
    elif args.gen_exclude_list:
        # 从 mc_cross_1.xlsx 生成排除列表
        if isinstance(__file__, bytes):
            file_path = __file__.decode('utf-8')
        else:
            file_path = __file__
        project_dir = os.path.dirname(os.path.dirname(os.path.abspath(file_path)))
        if isinstance(project_dir, bytes):
            project_dir = project_dir.decode('utf-8')
        xlsx_file = os.path.join(project_dir, u'data', u'test_data', u'mc_cross_1.xlsx')
        list_file = os.path.join(project_dir, u'data', u'test_data', u'mc_cross_1_exclude.list')
        _generate_exclude_list(xlsx_file, list_file)
    elif args.extract_cross_2:
        # 提取排除了mc_cross_1的剩余数据，生成mc_cross_2.xlsx
        print("[INFO] Generating mc_cross_2 (excluding mc_cross_1 data)...")
        extract_pattern_type_2_data(output_suffix='2')
    elif args.extract_cross_3:
        # 提取排除了mc_cross_1和mc_cross_2的剩余数据，生成mc_cross_3.xlsx
        print("[INFO] Generating mc_cross_3 (excluding mc_cross_1 and mc_cross_2 data)...")
        extract_pattern_type_2_data(output_suffix='3')
    elif args.extract_sc_cross_context:
        # 处理sc_cross.list文件，生成分类结果
        output_prefix = args.output_prefix if args.output_prefix else 'sc_cross'
        process_cross_list(input_file=args.input_file, output_prefix=output_prefix)
    elif args.extract_cross_context:
        # 处理mc_cross.list文件，生成分类结果
        output_prefix = args.output_prefix if args.output_prefix else 'mc_cross'
        process_cross_list(input_file=args.input_file, output_prefix=output_prefix)
    else:
        # 默认行为：处理mc_cross.list文件
        process_mc_cross_list()

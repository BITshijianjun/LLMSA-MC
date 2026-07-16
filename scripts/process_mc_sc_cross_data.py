# -*- coding: utf-8 -*-
from __future__ import print_function
from __future__ import unicode_literals
import os
import re
import io
import sys
import json
from collections import defaultdict

try:
    from openpyxl import Workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

def get_src_context_with_linenum(src_context, start_line):
    """Add a line number to each line of code."""
    updated_context = []
    for idx, ctx_line in enumerate(src_context):
        ctx_line = ctx_line.strip() + "//##" + str(start_line + idx)
        updated_context.append(ctx_line)
    return updated_context

def classify_line_type(line):
    """Classify the type based on the target line's code."""
    line = line.strip()
    if not line:
        return 'Other'

    # Return statement
    if line.startswith('return '):
        return 'Return Statement'

    # Loop statement
    if line.startswith('for (') or line.startswith('while ('):
        return 'Loop Statement'

    # struct/union/enum declaration
    if any(kw in line for kw in ['struct ', 'union ', 'enum ']) and '{' in line:
        return 'Type Declaration'

    # Variable declaration with assignment
    if re.match(r'(const\s+)?(struct|int|char|long|void|unsigned|bool|float|double)\s+\w+.*=', line):
        return 'Variable Declaration with Assignment'

    # Plain variable declaration
    if re.match(r'(const\s+)?(struct|int|char|long|void|unsigned|bool|float|double)\s+\w+', line):
        return 'Variable Declaration'

    # Assignment statement (not a declaration)
    if '=' in line and '==' not in line and not re.match(r'(const\s+)?(struct|int|char|long|void|unsigned|bool|float|double)', line):
        return 'Assignment Statement'

    # Function call with assignment
    if re.search(r'\w+\s*=\s*\w+\s*\([^)]*\)', line):
        return 'Function Call with Assignment'

    # Function call (no assignment)
    if re.search(r'^\s*\w+\s*\([^)]*\)', line):
        return 'Function Call'

    # Array/pointer access
    if '[' in line or ']' in line:
        return 'Array/Pointer Access'

    return 'Other'

def find_function_end_line(filepath, func_start_line):
    """Find the function's end line (the matching closing brace)."""
    try:
        with io.open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
            lines = f.readlines()

        if func_start_line < 1 or func_start_line > len(lines):
            return None

        # Starting from the function definition line, find the first {
        brace_count = 0
        found_opening = False

        for i in range(func_start_line - 1, len(lines)):
            line = lines[i]

            # Count braces
            for char in line:
                if char == '{':
                    brace_count += 1
                    found_opening = True
                elif char == '}':
                    if found_opening:
                        brace_count -= 1
                        if brace_count == 0:
                            return i + 1  # Return 1-based line number

        return None
    except:
        return None


def find_function_definition_line(filepath, func_name, start_from_line=None):
    """
    Find the line number of a function definition.
    If start_from_line is provided, search upward from that line (recommended).
    Otherwise search downward from the beginning of the file.
    """
    try:
        with io.open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
            lines = f.readlines()

        pattern = re.compile(r'\b' + re.escape(func_name) + r'\s*\(')

        # Pattern: identify a function definition (return type + function name + ( + possible parameters)
        def is_function_definition(line_idx):
            """Check whether this line is a function definition."""
            line = lines[line_idx]
            stripped = line.strip()

            # Exclude comment lines
            if stripped.startswith('//') or stripped.startswith('/*') or stripped.startswith('*'):
                return False

            # Check whether the function name and parenthesis are present
            if not pattern.search(line):
                return False

            # Check whether the current line has function-definition features (a return-type keyword)
            if any(kw in stripped for kw in ['void ', 'int ', 'bool ', 'char ', 'long ', 'struct ', 'static ', 'inline ', 'unsigned ', 'const ', '*']):
                return True

            # If the current line has no return type, check the preceding lines (handles multi-line function definitions)
            # Search upward up to 10 lines for a return-type keyword
            for i in range(max(0, line_idx - 10), line_idx):
                prev_line = lines[i].strip()
                if any(kw in prev_line for kw in ['int ', 'void ', 'bool ', 'char ', 'long ', 'struct ', 'static ', 'inline ', 'unsigned ', 'const ', '*']):
                    # Make sure there's no semicolon or other statement terminator between these lines
                    between_lines = '\n'.join(lines[i:line_idx+1])
                    if ';' not in between_lines.split('{')[0]:  # Ignore semicolons inside {}
                        return True

            return False

        # Strategy 1: if start_from_line is provided, search upward from there (recommended)
        if start_from_line is not None:
            start_idx = min(start_from_line - 1, len(lines) - 1)
            for i in range(start_idx, max(-1, start_idx - 500), -1):  # Search upward at most 500 lines
                if is_function_definition(i):
                    return i + 1

        # Strategy 2: search downward from the beginning of the file
        for lineno in range(len(lines)):
            if is_function_definition(lineno):
                return lineno + 1

        return None
    except:
        return None

def is_conditional_statement(line):
    """Check whether a line is a conditional statement (if, else if, else, while, for, etc.)."""
    line_stripped = line.strip()
    if not line_stripped:
        return False

    # Check whether the line starts with a conditional-statement keyword
    keywords = ['if ', 'else if ', 'else', 'while ', 'for ', 'switch ', 'case ']
    for kw in keywords:
        if line_stripped.startswith(kw):
            return True

    return False

def references_any_arg(line, args):
    """Check whether a line references any of the arguments."""
    if not args:
        return True

    for arg in args:
        # Use a word boundary to match the argument name
        if re.search(r'\b' + re.escape(arg) + r'\b', line):
            return True
    return False

def filter_lines_with_callsite_args(code_lines, callsite_code, start_line):
    """
    Filter the list of code lines down to lines that reference the callsite arguments.
    Logic:
    1. Keep the callsite line itself.
    2. Only drop conditional statements (if, else, while, for, etc.) that don't reference a callsite argument.
    3. Keep all other lines (regardless of whether they reference an argument).
    4. Add the original line number to each line.
    """
    if not callsite_code:
        return get_src_context_with_linenum(code_lines, start_line)

    # Extract arguments from the callsite code
    # Find the argument portion of the function call
    match = re.search(r'\w+\s*\(([^)]*)\)', callsite_code)
    if not match:
        return get_src_context_with_linenum(code_lines, start_line)

    args_str = match.group(1)
    if not args_str.strip():
        return get_src_context_with_linenum(code_lines, start_line)

    # Parse the argument list (simple split)
    args = [arg.strip() for arg in args_str.split(',')]
    # Keep only valid argument identifiers (drop empty strings and parts of complex expressions)
    args = [arg for arg in args if arg and not any(c in arg for c in ['(', ')', '[', ']'])]

    filtered_lines = []
    for idx, line in enumerate(code_lines):
        line_stripped = line.strip()
        original_line_num = start_line + idx

        # Keep the callsite line itself
        if callsite_code in line_stripped:
            filtered_lines.append(line_stripped + "//##" + str(original_line_num))
        # For conditional statements, only drop them when they don't reference an argument
        elif is_conditional_statement(line):
            if references_any_arg(line, args):
                # Conditional statement referencing an argument: keep it
                filtered_lines.append(line_stripped + "//##" + str(original_line_num))
            # Conditional statement not referencing an argument: drop it
        # For non-conditional statements, keep all of them
        else:
            filtered_lines.append(line_stripped + "//##" + str(original_line_num))

    return filtered_lines


def _generate_exclude_list(xlsx_file, output_list_file):
    """
    Generate an exclude list from an xlsx file (simple text format).
    Format: target_file|target_line|callsite_line
    """
    try:
        from openpyxl import load_workbook
        wb = load_workbook(xlsx_file)
        ws = wb.active

        with io.open(output_list_file, 'w', encoding='utf-8') as f:
            for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
                if row_idx == 1:  # Skip header
                    continue

                if row and len(row) >= 8:
                    try:
                        target_file = str(row[0]).strip() if row[0] else ''
                        target_line = str(row[1]).strip() if row[1] else ''
                        callsite_line = str(row[7]).strip() if row[7] else ''

                        if target_file and target_line and callsite_line:
                            line = u"{}|{}|{}\n".format(target_file, target_line, callsite_line)
                            f.write(line)
                    except:
                        pass

        return output_list_file
    except:
        return None


def _read_exclude_list(list_file):
    """
    Read records from the exclude list file.
    """
    records_set = set()

    if not os.path.exists(list_file):
        return records_set

    try:
        with io.open(list_file, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if line:
                    parts = line.split('|')
                    if len(parts) == 3:
                        record_id = tuple(parts)
                        records_set.add(record_id)

        print("[INFO] Loaded {} excluded records from list".format(len(records_set)))
        return records_set
    except:
        print("[WARNING] Failed to read exclude list")
        return records_set


def _read_xlsx_records(xlsx_file):
    """
    Read records from an Excel file and return a set of unique identifiers.
    Identifier: (target_func_filepath, target_line, callsite_line)
    """
    records_set = set()

    if not os.path.exists(xlsx_file):
        return records_set

    # Read directly with openpyxl; don't process error messages to avoid encoding issues
    try:
        from openpyxl import load_workbook
        wb = load_workbook(xlsx_file, data_only=True)
        ws = wb.active

        # Skip the header row
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            if row_idx == 1:  # Skip header
                continue

            if row and len(row) >= 8:
                try:
                    target_file = str(row[0]).strip() if row[0] else ''
                    target_line = str(row[1]).strip() if row[1] else ''
                    callsite_line = str(row[7]).strip() if row[7] else ''

                    if target_file and target_line and callsite_line:
                        record_id = (target_file, target_line, callsite_line)
                        records_set.add(record_id)
                except:
                    pass

        return records_set

    except:
        return records_set


def extract_pattern_type_2_data(output_suffix='1', exclude_xlsx=None, excluded_records=None):
    """
    Extract records with pattern_type == 2 from unchecked_locations_report.json,
    distributing them evenly across kernel subsystems, output as mc_cross_{suffix}.list format.

    Args:
        output_suffix: output file suffix (default '1', can also be '2', etc.)
        exclude_xlsx: path to the xlsx file to exclude (deprecated, use excluded_records instead)
        excluded_records: set of records to exclude (format: set of (target_file, target_line, callsite_line))
    """
    if isinstance(__file__, bytes):
        file_path = __file__.decode('utf-8')
    else:
        file_path = __file__
    project_dir = os.path.dirname(os.path.dirname(os.path.abspath(file_path)))
    if isinstance(project_dir, bytes):
        project_dir = project_dir.decode('utf-8')
    json_file = os.path.join(project_dir, u'scripts', u'unchecked_locations_report.json')
    output_file = os.path.join(project_dir, u'data', u'test_data', u'mc_cross_{}.xlsx'.format(output_suffix))
    temp_output_file = output_file + u'.tmp'
    target_count = 100

    # Determine the set of records to exclude
    excluded_records_set = set()

    # Determine which files to exclude based on output_suffix
    if output_suffix == '2':
        # mc_cross_2: exclude mc_cross_1
        selected_files = [u'mc_cross_1_selected.txt']
        exclude_label = 'mc_cross_1'
    elif output_suffix == '3':
        # mc_cross_3: exclude mc_cross_1 and mc_cross_2
        selected_files = [u'mc_cross_1_selected.txt', u'mc_cross_2_selected.txt']
        exclude_label = 'mc_cross_1 and mc_cross_2'
    else:
        selected_files = []
        exclude_label = None

    # Read the selected-records file
    if selected_files:
        for selected_file_name in selected_files:
            selected_records_file = os.path.join(project_dir, u'data', u'test_data', selected_file_name)

            # If the selected-records file doesn't exist, try generating it from the corresponding xlsx
            if not os.path.exists(selected_records_file):
                # Infer the xlsx filename from the file name
                xlsx_suffix = selected_file_name.split('_')[2].split('.')[0]  # Extract X from mc_cross_X_selected.txt
                mc_cross_xlsx = os.path.join(project_dir, u'data', u'test_data', u'mc_cross_{}.xlsx'.format(xlsx_suffix))
                if os.path.exists(mc_cross_xlsx):
                    _generate_exclude_list(mc_cross_xlsx, selected_records_file)

            # Try reading the selected-records file
            if os.path.exists(selected_records_file):
                try:
                    with io.open(selected_records_file, 'r', encoding='utf-8') as f:
                        for line in f:
                            line = line.strip()
                            if line:
                                parts = line.split('|')
                                if len(parts) == 3:
                                    excluded_records_set.add(tuple(parts))
                except:
                    pass

        if excluded_records_set:
            try:
                print("[INFO] Loaded {} records to exclude from {}".format(len(excluded_records_set), exclude_label))
            except:
                pass

    if excluded_records:
        excluded_records_set = excluded_records
    elif exclude_xlsx:
        excluded_records_set = _read_xlsx_records(exclude_xlsx)

    try:
        print("[EXTRACT] Reading {}".format(json_file))
    except:
        print("[EXTRACT] Reading JSON file...")
    if not os.path.exists(json_file):
        try:
            print("[ERROR] File not found: {}".format(json_file))
        except:
            print("[ERROR] File not found")
        return

    # Read the JSON file
    with io.open(json_file, 'r', encoding='utf-8') as f:
        all_data = json.load(f)

    # Filter by pattern_type == 2 and group by subsystem and file
    data_by_subsystem_file = defaultdict(lambda: defaultdict(list))
    pattern_type_2_data = [d for d in all_data if d.get('pattern_type') == 2]

    print("[EXTRACT] Found {} records with pattern_type=2".format(len(pattern_type_2_data)))

    # If an exclude set is provided, filter out records contained in it
    if excluded_records_set:
        original_count = len(pattern_type_2_data)
        pattern_type_2_data = [
            d for d in pattern_type_2_data
            if (d.get('target_func_filepath', ''), str(d.get('target_line', '')), str(d.get('callsite_line', '')))
            not in excluded_records_set
        ]
        print("[EXTRACT] After excluding: {} records (filtered out {} duplicates)".format(
            len(pattern_type_2_data), original_count - len(pattern_type_2_data)))

    for item in pattern_type_2_data:
        subsystem = item.get('subsystem', 'unknown')
        filepath = item.get('filepath', 'unknown')
        data_by_subsystem_file[subsystem][filepath].append(item)

    print("[EXTRACT] Distribution by subsystem and file:")
    subsystems = sorted(data_by_subsystem_file.keys())
    for subsystem in subsystems:
        file_count = len(data_by_subsystem_file[subsystem])
        total_records = sum(len(items) for items in data_by_subsystem_file[subsystem].values())
        print("  {}: {} files, {} records".format(subsystem, file_count, total_records))

    # Global sampling: round-robin across subsystems and files, ensuring the total is exactly target_count
    selected_data = []
    global_samples_per_file = defaultdict(int)
    subsystem_file_index = {}

    print("\n[EXTRACT] Extracting {} records (distributing across subsystems and files):".format(target_count))
    print("[DEBUG] Subsystems (sorted): {}".format(subsystems))

    # Save the list of selected records for use by the subsequent mc_cross_N+1
    selected_records_file = os.path.join(project_dir, u'data', u'test_data', u'mc_cross_{}_selected.txt'.format(output_suffix))

    # Initialize each subsystem's file list and sampling position
    for subsystem in subsystems:
        files = sorted(data_by_subsystem_file[subsystem].keys())
        subsystem_file_index[subsystem] = {f: 0 for f in files}

    # Round-robin sampling: iterate over subsystems and files until the target count is reached or all records are exhausted
    subsystem_idx = 0
    all_exhausted = False  # Flag: all subsystems' records have been exhausted

    while len(selected_data) < target_count and not all_exhausted:
        if subsystem_idx >= len(subsystems):
            subsystem_idx = 0

        subsystem = subsystems[subsystem_idx]
        files_in_subsystem = sorted(data_by_subsystem_file[subsystem].keys())

        # Round-robin over files within this subsystem
        file_found = False
        for _ in range(len(files_in_subsystem)):
            # Find a file that still has samples available
            for filepath in files_in_subsystem:
                records_in_file = data_by_subsystem_file[subsystem][filepath]
                current_idx = subsystem_file_index[subsystem][filepath]

                if current_idx < len(records_in_file):
                    # Extract one sample from this file
                    record = records_in_file[current_idx]
                    selected_data.append(record)
                    subsystem_file_index[subsystem][filepath] += 1
                    global_samples_per_file[filepath] += 1
                    file_found = True
                    break

            if file_found or len(selected_data) >= target_count:
                break

        subsystem_idx += 1

        # Check whether all subsystems' records have been exhausted
        all_exhausted = all(
            all(subsystem_file_index[sub][f] >= len(data_by_subsystem_file[sub][f])
                for f in sorted(data_by_subsystem_file[sub].keys()))
            for sub in subsystems
        )

        if len(selected_data) >= target_count:
            break

    # Truncate to the target count
    selected_data = selected_data[:target_count]

    # Warn if the actual number of records obtained is less than the target count
    if len(selected_data) < target_count:
        try:
            print("[WARNING] Only {} records available (target was {}), using all available records".format(
                len(selected_data), target_count))
        except:
            pass

    # Generate a sampling summary for verification (to ensure consistency across runs)
    try:
        sample_count = min(5, len(selected_data))
        print("[DEBUG] Sample verification (first {} records):".format(sample_count))
        for idx in range(sample_count):
            record = selected_data[idx]
            subsystem = record.get('subsystem', 'unknown')
            try:
                print("  Record {}: {}".format(idx + 1, subsystem))
            except:
                print("  Record {}".format(idx + 1))
    except:
        pass

    # Tally the results
    subsystem_counts = defaultdict(int)
    subsystem_files = defaultdict(set)
    for record in selected_data:
        subsystem = record.get('subsystem', 'unknown')
        subsystem_counts[subsystem] += 1
        filepath = record.get('filepath', 'unknown')
        subsystem_files[subsystem].add(filepath)

    for subsystem in sorted(subsystem_counts.keys()):
        print("  {}: {} records from {} files".format(
            subsystem, subsystem_counts[subsystem], len(subsystem_files[subsystem])))

    print("[EXTRACT] Total selected: {} records".format(len(selected_data)))

    # When generating mc_cross_1, save the selected records to a file for mc_cross_2 to use
    if selected_records_file:
        try:
            with io.open(selected_records_file, 'w', encoding='utf-8') as f:
                for record in selected_data:
                    target_file = record.get('target_func_filepath', '')
                    target_line = str(record.get('target_line', ''))
                    callsite_line = str(record.get('callsite_line', ''))
                    if target_file and target_line and callsite_line:
                        line = u"{}|{}|{}\n".format(target_file, target_line, callsite_line)
                        f.write(line)
        except:
            pass

    # Convert the path format (extract the relative portion from the full path)
    def extract_relative_path(filepath):
        """Extract the linux-* relative portion from the full path."""
        # Find the position of "linux-"
        match = re.search(r'(linux-[\w\./-]+)', filepath)
        if match:
            return match.group(1).replace('\\', '/')
        return filepath

    # Output in Excel format or txt format
    try:
        print("[EXTRACT] Writing to output file...")
    except:
        print("[EXTRACT] Writing to output file...")

    valid_records = 0

    # Try generating an Excel file using openpyxl
    if OPENPYXL_AVAILABLE:
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = u'mc_cross'

            # Write the header
            headers = [
                u'Target File Path',
                u'Targeted Line Number',
                u'Targeted Line Code',
                u'Target Function',
                u'Caller Function',
                u'Callsite',
                u'Caller File Path',
                u'Callsite Line Number'
            ]
            ws.append(headers)

            # Write the data
            for item in selected_data:
                target_file = item.get('target_func_filepath', '')
                target_line = str(item.get('target_line', ''))
                target_code = str(item.get('target_statement', ''))
                target_func = str(item.get('target_func', ''))
                caller_func = str(item.get('caller_func', ''))
                callsite = str(item.get('callsite_statement', ''))
                caller_file = item.get('filepath', '')
                callsite_line = str(item.get('callsite_line', ''))

                # Only write records with valid paths
                if target_file.strip() and caller_file.strip():
                    ws.append([
                        target_file,
                        target_line,
                        target_code,
                        target_func,
                        caller_func,
                        callsite,
                        caller_file,
                        callsite_line
                    ])
                    valid_records += 1

            # Adjust column widths
            ws.column_dimensions['A'].width = 30  # Target File Path
            ws.column_dimensions['B'].width = 15  # Targeted Line Number
            ws.column_dimensions['C'].width = 25  # Targeted Line Code
            ws.column_dimensions['D'].width = 20  # Target Function
            ws.column_dimensions['E'].width = 20  # Caller Function
            ws.column_dimensions['F'].width = 30  # Callsite
            ws.column_dimensions['G'].width = 30  # Caller File Path
            ws.column_dimensions['H'].width = 18  # Callsite Line Number

            # Save the file to a temporary location, then rename it
            wb.save(temp_output_file)
            # Delete the old file and rename
            try:
                if os.path.exists(output_file):
                    os.remove(output_file)
            except:
                pass
            os.rename(temp_output_file, output_file)

        except Exception as e:
            try:
                print("[WARNING] Failed to save as Excel: {}".format(str(e)))
            except:
                print("[WARNING] Failed to save as Excel")
            # Fall back to txt format
            _save_as_txt = True
    else:
        _save_as_txt = True

    # If Excel isn't used or saving as Excel failed, save as txt format
    if '_save_as_txt' in locals() and _save_as_txt:
        output_txt = output_file.replace(u'.xlsx', u'.list')
        try:
            print("[EXTRACT] Saving as text format instead...")
        except:
            pass

        with io.open(output_txt, 'w', encoding='utf-8') as f:
            # Write the header
            header = u"Target File Path\tTargeted Line Number\tTargeted Line Code\tTarget Function\tCaller Function\tCallsite\tCaller File Path\tCallsite Line Number\n"
            f.write(header)

            # Write the data
            valid_records = 0
            for item in selected_data:
                target_file = item.get('target_func_filepath', '')
                target_line = str(item.get('target_line', ''))
                target_code = str(item.get('target_statement', ''))
                target_func = str(item.get('target_func', ''))
                caller_func = str(item.get('caller_func', ''))
                callsite = str(item.get('callsite_statement', ''))
                caller_file = item.get('filepath', '')
                callsite_line = str(item.get('callsite_line', ''))

                # Only write records with valid paths
                if target_file.strip() and caller_file.strip():
                    line = u"{}\t{}\t{}\t{}\t{}\t{}\t{}\t{}\n".format(
                        target_file,
                        target_line,
                        target_code,
                        target_func,
                        caller_func,
                        callsite,
                        caller_file,
                        callsite_line
                    )
                    f.write(line)
                    valid_records += 1
        output_file = output_txt

    try:
        print("[SUCCESS] Extracted {} records to {} ({} valid records)".format(len(selected_data), output_file, valid_records))
    except:
        print("[SUCCESS] Extracted {} records ({} valid records)".format(len(selected_data), valid_records))

    # Output statistics by subsystem
    print("\n[SUMMARY] Distribution in output:")
    subsystem_dist = defaultdict(int)
    for item in selected_data:
        subsystem = item.get('subsystem', 'unknown')
        subsystem_dist[subsystem] += 1

    for subsystem in sorted(subsystem_dist.keys()):
        print("  {}: {} records".format(subsystem, subsystem_dist[subsystem]))


def process_cross_list(input_file=None, output_prefix='mc_cross'):
    """
    Generic function for processing a cross.list file.

    Args:
        input_file: input file path (if None, uses the default mc_cross.list or sc_cross.list)
        output_prefix: prefix for the output files (default 'mc_cross', can also be 'sc_cross', etc.)
    """
    if isinstance(__file__, bytes):
        file_path = __file__.decode('utf-8')
    else:
        file_path = __file__
    project_dir = os.path.dirname(os.path.dirname(os.path.abspath(file_path)))
    if isinstance(project_dir, bytes):
        project_dir = project_dir.decode('utf-8')

    # Determine the input file path
    if input_file is None:
        # If no input file is specified, determine it based on output_prefix
        # If output_prefix contains 'sc', use sc_cross.list; otherwise use mc_cross.list
        if 'sc' in output_prefix.lower():
            default_input_name = 'sc_cross.list'
        else:
            default_input_name = 'mc_cross.list'
        input_file = os.path.join(project_dir, u'data', u'test_data', default_input_name)
    elif not os.path.isabs(input_file):
        # If it's a relative path, resolve it relative to project_dir
        input_file = os.path.join(project_dir, input_file)

    output_full_ctx = os.path.join(project_dir, u'data', u'test_data', u'{}_classify_full_contexts.list'.format(output_prefix))
    output_classify = os.path.join(project_dir, u'data', u'test_data', u'{}_classify.list'.format(output_prefix))
    kernel_code_dir = os.path.join(project_dir, u'data', u'kernel-code')

    results_full_ctx = []
    results_classify = []

    # Extract the input filename for display
    try:
        input_filename = os.path.basename(input_file)
        print("[START] Processing {} with Target Function".format(input_filename))
        print("[INFO] Input file: {}".format(input_file))
        print("[INFO] Output prefix: {}".format(output_prefix))
    except:
        # Handle encoding issues
        print("[START] Processing cross list file with Target Function")
        print("[INFO] Output prefix: {}".format(output_prefix))

    if not os.path.exists(input_file):
        try:
            print("[ERROR] Input file not found: {}".format(input_file))
        except:
            print("[ERROR] Input file not found (path encoding issue)")
        return

    with io.open(input_file, 'r', encoding='utf-8') as f:
        lines = f.readlines()[1:]  # Skip header

    # Deduplication logic: detect and remove duplicate lines
    seen_lines = {}  # {line_content: first_occurrence_idx}
    duplicate_samples = []  # Records the sample indices that are duplicates
    unique_lines = []  # Holds the deduplicated lines

    for idx, line in enumerate(lines):
        if not line.strip():
            continue

        line_content = line.strip()
        if line_content in seen_lines:
            # Found a duplicate line
            first_idx = seen_lines[line_content]
            duplicate_samples.append((idx + 1, first_idx + 1))  # 1-based sample number
        else:
            # First time seeing this line
            seen_lines[line_content] = idx
            unique_lines.append((idx, line))

    # If there are duplicates, print them
    if duplicate_samples:
        print("[WARNING] Found {} duplicate records:".format(len(duplicate_samples)))
        for dup_idx, first_idx in duplicate_samples:
            print("  Sample #{} is a duplicate of Sample #{}".format(dup_idx, first_idx))

    print("Processing {} records (after removing {} duplicates)".format(len(unique_lines), len(duplicate_samples)))

    for original_idx, line in unique_lines:
        parts = line.strip().split('\t')
        if len(parts) < 8:
            continue

        # Parse the input fields
        target_file = parts[0]
        target_line_num = int(parts[1])
        target_code = parts[2]
        target_func = parts[3]
        caller_func = parts[4]
        callsite_code = parts[5]
        caller_file = parts[6]
        callsite_line_num = int(parts[7])

        # Classify based on the target line code (not the callsite code)
        line_type = classify_line_type(target_code)

        # Currently all lines default to "Yes" for Target Parameter Annotation
        param_annotation = "Yes"

        # Build the full path
        try:
            target_file_path = os.path.join(kernel_code_dir, target_file.replace('/', os.sep))
        except:
            target_file_path = kernel_code_dir + '\\' + target_file.replace('/', '\\')
        try:
            caller_file_path = os.path.join(kernel_code_dir, caller_file.replace('/', os.sep))
        except:
            caller_file_path = kernel_code_dir + '\\' + caller_file.replace('/', '\\')

        # Check whether the files exist
        if not os.path.exists(target_file_path) or not os.path.exists(caller_file_path):
            continue


        # Extract the context
        try:
            # Read both files
            try:
                with io.open(caller_file_path, 'r', encoding='utf-8', errors='ignore') as f:
                    caller_content = f.readlines()
            except Exception as read_err:
                print("[ERROR] Sample #{}: Failed to read caller file: {}".format(original_idx + 1, str(read_err)))
                raise

            try:
                with io.open(target_file_path, 'r', encoding='utf-8', errors='ignore') as f:
                    target_content = f.readlines()
            except Exception as read_err:
                print("[ERROR] Sample #{}: Failed to read target file: {}".format(original_idx + 1, str(read_err)))
                raise

            context_list = []

            # 1. Find the caller function's definition line (search upward from the callsite)
            caller_func_def_line = find_function_definition_line(caller_file_path, caller_func, callsite_line_num)

            # 2. Find the target function's definition line (search upward from target_line)
            target_func_def_line = find_function_definition_line(target_file_path, target_func, target_line_num)

            # Initialize the context lists
            full_context_list = []
            classify_context_list = []

            # Debug: print a warning if the function definition isn't found
            if not caller_func_def_line:
                try:
                    print("[WARNING] Sample #{}: Cannot find caller function '{}' in {}".format(
                        original_idx + 1, caller_func, os.path.basename(caller_file_path)))
                except:
                    pass
            if not target_func_def_line:
                try:
                    print("[WARNING] Sample #{}: Cannot find target function '{}' in {}".format(
                        original_idx + 1, target_func, os.path.basename(target_file_path)))
                except:
                    pass

            # If both lines were found, extract the context
            if caller_func_def_line and target_func_def_line:
                # ==================== full_contexts: caller function entry to the callsite + target function entry to the target line ====================
                # Extract from the caller function entry to the callsite
                full_caller_start = caller_func_def_line - 1
                full_caller_end = callsite_line_num
                caller_part_full = caller_content[full_caller_start:full_caller_end]
                caller_lines_full = get_src_context_with_linenum(caller_part_full, full_caller_start + 1)

                # Extract from the target function entry to the target line
                full_target_start = target_func_def_line - 1
                full_target_end = target_line_num
                target_part_full = target_content[full_target_start:full_target_end]
                target_lines_full = get_src_context_with_linenum(target_part_full, full_target_start + 1)

                full_context_list = caller_lines_full + target_lines_full

                # ==================== classify_contexts: select range based on param_annotation ====================
                if param_annotation == "Yes":
                    # From the caller function entry to the callsite, filtering for lines that reference a callsite argument
                    classify_caller_start = caller_func_def_line - 1
                    classify_caller_end = callsite_line_num
                    caller_part_classify = caller_content[classify_caller_start:classify_caller_end]
                    caller_lines_classify = filter_lines_with_callsite_args(caller_part_classify, callsite_code, classify_caller_start + 1)

                    # From the target function entry to the target line, keeping all lines
                    classify_target_start = target_func_def_line - 1
                    classify_target_end = target_line_num
                    target_part_classify = target_content[classify_target_start:classify_target_end]
                    target_lines_classify = get_src_context_with_linenum(target_part_classify, classify_target_start + 1)

                    classify_context_list = caller_lines_classify + target_lines_classify

            results_full_ctx.append({
                'file': target_file,
                'line': parts[1],
                'targeted_line_code': target_code,
                'caller_function': caller_func,
                'callsite': callsite_code,
                'context': full_context_list,
                'type': line_type,
                'param_annotation': param_annotation
            })

            results_classify.append({
                'file': target_file,
                'line': parts[1],
                'targeted_line_code': target_code,
                'caller_function': caller_func,
                'callsite': callsite_code,
                'context': classify_context_list,
                'type': line_type,
                'param_annotation': param_annotation
            })

        except Exception as e:
            pass

    # Save the results
    print("\n[SAVE] {} full_context records".format(len(results_full_ctx)))
    with io.open(output_full_ctx, 'w', encoding='utf-8') as f:
        for r in results_full_ctx:
            ctx_str = str(r['context'])
            line = "{} ;; {} ;; {} ;; {} ;; {} ;; {} ;; {} ;; {}\n".format(
                '../data/kernel-code/' + r['file'],
                r['line'],
                r['targeted_line_code'],
                r['caller_function'],
                r['callsite'],
                ctx_str,
                r['type'],
                r['param_annotation']
            )
            f.write(line)

    print("[SAVE] {} classify records".format(len(results_classify)))
    with io.open(output_classify, 'w', encoding='utf-8') as f:
        for r in results_classify:
            ctx_str = str(r['context'])
            line = "{} ;; {} ;; {} ;; {} ;; {} ;; {} ;; {} ;; {}\n".format(
                '../data/kernel-code/' + r['file'],
                r['line'],
                r['targeted_line_code'],
                r['caller_function'],
                r['callsite'],
                ctx_str,
                r['type'],
                r['param_annotation']
            )
            f.write(line)

    print("\n[SUMMARY]")
    dist = {}
    for r in results_classify:
        t = r['type']
        dist[t] = dist.get(t, 0) + 1
    print("  Total: {}".format(len(results_classify)))
    for t, c in sorted(dist.items()):
        print("    {}: {}".format(t, c))


def process_mc_cross_list():
    """Process the mc_cross.list file."""
    process_cross_list(input_file=None, output_prefix='mc_cross')


def process_sc_cross_list():
    """Process the sc_cross.list file."""
    process_cross_list(input_file=None, output_prefix='sc_cross')


# Handle command-line arguments
if __name__ == '__main__':
    import argparse

    parser = argparse.ArgumentParser(description='Process cross.list files and extract context')
    parser.add_argument('--extract-cross', action='store_true',
                        help='Extract pattern_type=2 data from unchecked_locations_report.json (mc_cross_1)')
    parser.add_argument('--extract-cross-2', action='store_true',
                        help='Extract remaining pattern_type=2 data (excluding mc_cross_1, generate mc_cross_2)')
    parser.add_argument('--extract-cross-3', action='store_true',
                        help='Extract remaining pattern_type=2 data (excluding mc_cross_1 and mc_cross_2, generate mc_cross_3)')
    parser.add_argument('--gen-exclude-list', action='store_true',
                        help='Generate exclude list from mc_cross_1.xlsx for mc_cross_2 extraction')
    parser.add_argument('--extract-cross-context', action='store_true',
                        help='Process mc_cross.list and generate classify results')
    parser.add_argument('--extract-sc-cross-context', action='store_true',
                        help='Process sc_cross.list and generate classify results')
    parser.add_argument('--input-file', type=str, default=None,
                        help='Custom input file path (absolute or relative to project directory)')
    parser.add_argument('--output-prefix', type=str, default=None,
                        help='Output file prefix (default: depends on the data type)')

    args = parser.parse_args()

    if args.extract_cross:
        # Extract pattern_type=2 data from unchecked_locations_report.json, generating mc_cross_1.xlsx
        extract_pattern_type_2_data(output_suffix='1')
    elif args.gen_exclude_list:
        # Generate the exclude list from mc_cross_1.xlsx
        if isinstance(__file__, bytes):
            file_path = __file__.decode('utf-8')
        else:
            file_path = __file__
        project_dir = os.path.dirname(os.path.dirname(os.path.abspath(file_path)))
        if isinstance(project_dir, bytes):
            project_dir = project_dir.decode('utf-8')
        xlsx_file = os.path.join(project_dir, u'data', u'test_data', u'mc_cross_1.xlsx')
        list_file = os.path.join(project_dir, u'data', u'test_data', u'mc_cross_1_exclude.list')
        _generate_exclude_list(xlsx_file, list_file)
    elif args.extract_cross_2:
        # Extract the remaining data excluding mc_cross_1, generating mc_cross_2.xlsx
        print("[INFO] Generating mc_cross_2 (excluding mc_cross_1 data)...")
        extract_pattern_type_2_data(output_suffix='2')
    elif args.extract_cross_3:
        # Extract the remaining data excluding mc_cross_1 and mc_cross_2, generating mc_cross_3.xlsx
        print("[INFO] Generating mc_cross_3 (excluding mc_cross_1 and mc_cross_2 data)...")
        extract_pattern_type_2_data(output_suffix='3')
    elif args.extract_sc_cross_context:
        # Process the sc_cross.list file and generate classification results
        output_prefix = args.output_prefix if args.output_prefix else 'sc_cross'
        process_cross_list(input_file=args.input_file, output_prefix=output_prefix)
    elif args.extract_cross_context:
        # Process the mc_cross.list file and generate classification results
        output_prefix = args.output_prefix if args.output_prefix else 'mc_cross'
        process_cross_list(input_file=args.input_file, output_prefix=output_prefix)
    else:
        # Default behavior: process the mc_cross.list file
        process_mc_cross_list()
